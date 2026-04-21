"""
Project 7: Excel/CSV Dashboard Generator
Author: mikey811
Description: Reads cleaned CSVs and auto-generates a styled multi-tab Excel
             workbook with summary tables, charts, and conditional formatting.
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# CONFIG
INPUT_DIR = "data"
OUTPUT_FILE = "dashboard.xlsx"

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL    = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN        = Side(style="thin", color="AAAAAA")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row_idx, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.border    = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_data_rows(ws, start_row, end_row, num_cols):
    for r in range(start_row, end_row + 1):
        fill = ALT_FILL if r % 2 == 0 else None
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            if fill:
                cell.fill = fill
            cell.border    = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")


def auto_column_widths(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def df_to_sheet(wb, df, sheet_name):
    ws = wb.create_sheet(title=sheet_name[:31])
    ws.row_dimensions[1].height = 20
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
    style_header_row(ws, 1, len(df.columns))
    style_data_rows(ws, 2, len(df) + 1, len(df.columns))
    auto_column_widths(ws)
    ws.freeze_panes = "A2"
    return ws


def build_summary_sheet(wb, file_summary):
    ws = wb.create_sheet(title="Summary", index=0)
    ws.sheet_view.showGridLines = False
    title_cell = ws["B2"]
    title_cell.value = "Data Dashboard - Auto-Generated Report"
    title_cell.font  = Font(size=16, bold=True, color="1F4E79")
    ws.merge_cells("B2:F2")
    ts_cell = ws["B3"]
    ts_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ts_cell.font  = Font(size=10, italic=True, color="888888")
    ws.merge_cells("B3:F3")
    headers = ["File", "Rows", "Columns", "Sheet Tab", "Null Values"]
    for c_idx, h in enumerate(headers, 2):
        cell = ws.cell(row=5, column=c_idx, value=h)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.border    = BORDER
        cell.alignment = Alignment(horizontal="center")
    for r_idx, entry in enumerate(file_summary, 6):
        row_data = [entry["file"], entry["rows"], entry["cols"], entry["sheet"], entry["nulls"]]
        for c_idx, val in enumerate(row_data, 2):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            if r_idx % 2 == 0:
                cell.fill = ALT_FILL
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 22


def build_chart_sheet(wb, file_summary):
    ws = wb.create_sheet(title="Row Counts", index=1)
    ws["A1"] = "File"
    ws["B1"] = "Row Count"
    for i, entry in enumerate(file_summary, 2):
        ws.cell(row=i, column=1, value=entry["file"])
        ws.cell(row=i, column=2, value=entry["rows"])
    chart = BarChart()
    chart.type         = "col"
    chart.title        = "Row Count by Source File"
    chart.y_axis.title = "Rows"
    chart.x_axis.title = "File"
    chart.style        = 10
    chart.width        = 20
    chart.height       = 12
    data = Reference(ws, min_col=2, min_row=1, max_row=len(file_summary) + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(file_summary) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D2")


def main():
    if not os.path.isdir(INPUT_DIR):
        os.makedirs(INPUT_DIR)
        sample = pd.DataFrame({
            "date":     ["2024-01-01", "2024-01-02", "2024-01-03"],
            "product":  ["Widget A",   "Widget B",   "Widget A"],
            "quantity": [120, 85, 200],
            "revenue":  [2400.0, 1700.0, 4000.0],
        })
        sample.to_csv(os.path.join(INPUT_DIR, "sample_sales.csv"), index=False)
        print("[INFO] No data/ folder found - created sample_sales.csv for demo.")

    csv_files = [f for f in os.listdir(INPUT_DIR) if f.endswith(".csv")]
    if not csv_files:
        print("[WARN] No CSV files found in data/. Exiting.")
        return

    wb = Workbook()
    wb.remove(wb.active)
    file_summary = []

    for fname in sorted(csv_files):
        path = os.path.join(INPUT_DIR, fname)
        try:
            df = pd.read_csv(path)
        except Exception as e:
            print(f"[ERROR] Could not read {fname}: {e}")
            continue
        sheet_name = os.path.splitext(fname)[0].replace("_", " ").title()[:31]
        df_to_sheet(wb, df, sheet_name)
        file_summary.append({
            "file":  fname,
            "rows":  len(df),
            "cols":  len(df.columns),
            "sheet": sheet_name,
            "nulls": int(df.isnull().sum().sum()),
        })
        print(f"[OK] {fname} -> '{sheet_name}' ({len(df)} rows)")

    build_summary_sheet(wb, file_summary)
    build_chart_sheet(wb, file_summary)
    wb.save(OUTPUT_FILE)
    print(f"[DONE] Dashboard saved to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
